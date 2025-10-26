import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pathlib import Path


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    try:
        csv_path_obj = Path(csv_path)
        if not csv_path_obj.exists():
            raise FileNotFoundError(f"CSV файл не найден: {csv_path}")

        xlsx_dir = Path(xlsx_path).parent
        xlsx_dir.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        with open(csv_path, 'r', encoding='utf-8') as csv_file:
            reader = csv.reader(csv_file)
            for row in reader:
                ws.append(row)

        if ws.max_row == 0:
            raise ValueError("CSV файл пуст")

        for col in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            
            for row in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            adjusted_width = max(max_length + 2, 8)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(xlsx_path)
        print(f"✓ Успешно создан XLSX файл: {xlsx_path}")
        
    except FileNotFoundError as e:
        raise e
    except ValueError as e:
        raise e
    except Exception as e:
        raise ValueError(f"Ошибка при преобразовании CSV в XLSX: {e}")


if __name__ == "__main__":
    try:
        print("=== Тестирование CSV to XLSX ===")

        current_dir = Path(__file__).parent
        samples_dir = current_dir / "../../data/samples"
        out_dir = current_dir / "../../data/out"

        out_dir.mkdir(parents=True, exist_ok=True)

        print("1. Конвертация people.csv...")
        csv_to_xlsx(
            str(samples_dir / "people.csv"),
            str(out_dir / "people.xlsx")
        )

        cities_csv = samples_dir / "cities.csv"
        if cities_csv.exists():
            print("2. Конвертация cities.csv...")
            csv_to_xlsx(
                str(cities_csv),
                str(out_dir / "cities.xlsx")
            )
        else:
            print("2. cities.csv не найден, пропускаем")
            
        print("✓ Все операции завершены успешно!")
        
    except Exception as e:
        print(f"❌ Ошибка: {e}")